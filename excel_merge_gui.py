import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from datetime import datetime
import os

try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

class ToolTip(object):
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tipwindow = None
        # Only show tooltip on mouse hover, not on focus/tab
        widget.bind("<Enter>", self.show_tip)
        widget.bind("<Leave>", self.hide_tip)
        # Do NOT bind to <FocusIn> or <FocusOut>

    def show_tip(self, event=None):
        if self.tipwindow or not self.text:
            return
        # Only show on mouse events
        if event is not None and event.type != tk.EventType.Enter:
            return
        x, y, cx, cy = self.widget.bbox("insert") if self.widget.winfo_ismapped() else (0,0,0,0)
        x = x + self.widget.winfo_rootx() + 25
        y = y + self.widget.winfo_rooty() + 20
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, justify=tk.LEFT,
                         background="#ffffe0", relief=tk.SOLID, borderwidth=1,
                         font=("tahoma", "9", "normal"))
        label.pack(ipadx=6, ipady=2)

    def hide_tip(self, event=None):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

class ExcelMergeApp:


    def __init__(self, master):
        self.master = master
        master.title("Excel Sheet Merge Tool")
        master.geometry("570x600")
        master.configure(bg="#f4f6fb")
        master.resizable(False, False)

        self.file1_path = tk.StringVar()
        self.datetime_str = tk.StringVar()

        label_font = ("Segoe UI", 11)
        entry_font = ("Segoe UI", 11)
        button_font = ("Segoe UI", 11, "bold")
        heading_font = ("Segoe UI", 16, "bold")
        section_font = ("Segoe UI", 12, "bold")

        # Logo area
        logo_frame = tk.Frame(master, bg="#f4f6fb")
        logo_frame.pack(pady=(10, 0))
        if PIL_AVAILABLE:
            try:
                logo_img = Image.open(os.path.join(os.path.dirname(__file__), "excel_logo.png"))
                logo_img = logo_img.resize((40, 40), Image.LANCZOS)
                self.logo_photo = ImageTk.PhotoImage(logo_img)
                tk.Label(logo_frame, image=self.logo_photo, bg="#f4f6fb").pack()
            except Exception:
                tk.Label(logo_frame, text="📊", font=("Segoe UI Emoji", 32), bg="#f4f6fb").pack()
        else:
            tk.Label(logo_frame, text="📊", font=("Segoe UI Emoji", 32), bg="#f4f6fb").pack()

        # Title
        tk.Label(master, text="Excel Sheet Merge Tool", font=heading_font, bg="#f4f6fb", fg="#2b3a67").pack(pady=(0, 3))
        tk.Label(master, text="Easily merge and update Excel files based on date-time.", font=("Segoe UI", 10), bg="#f4f6fb", fg="#4a4a4a").pack(pady=(0, 10))

        # Frame for file selectors
        file_frame = tk.Frame(master, bg="#f4f6fb", highlightbackground="#d1d9e6", highlightthickness=1, bd=0)
        file_frame.pack(pady=5, fill=tk.X, padx=30)

        # File 1
        tk.Label(file_frame, text="File 1 (where changes happen):", font=section_font, bg="#f4f6fb", anchor="w").grid(row=0, column=0, sticky="w", pady=(0,2))
        file1_entry = tk.Entry(file_frame, textvariable=self.file1_path, width=36, font=entry_font, relief="groove", bd=2)
        file1_entry.grid(row=1, column=0, padx=(0,8), sticky="w")
        file1_btn = tk.Button(file_frame, text="Browse", command=self.browse_file1, font=button_font, bg="#5b9bd5", fg="white", activebackground="#3d6fa5", relief="flat", cursor="hand2")
        file1_btn.grid(row=1, column=1, sticky="w")
        ToolTip(file1_btn, "Select the Excel file you want to update.")


        # Date-time input
        dt_frame = tk.Frame(master, bg="#f4f6fb")
        dt_frame.pack(pady=(18, 0), fill=tk.X, padx=30)
        tk.Label(dt_frame, text="Enter Date-Time (YYYY-MM-DD HH:MM):", font=label_font, bg="#f4f6fb").grid(row=0, column=0, sticky="w")
        dt_entry = tk.Entry(dt_frame, textvariable=self.datetime_str, width=20, font=entry_font, relief="groove", bd=2)
        dt_entry.grid(row=0, column=1, padx=(12,0))
        ToolTip(dt_entry, "Type the date and time (e.g. 2025-07-18 10:15)")



        # Buttons frame
        btns_frame = tk.Frame(master, bg="#f4f6fb")
        btns_frame.pack(pady=25)

        execute_btn = tk.Button(btns_frame, text="Execute", command=self.process_files, font=button_font, bg="#2b3a67", fg="white", activebackground="#1a2240", relief="flat", cursor="hand2")
        execute_btn.grid(row=0, column=0, padx=10)
        ToolTip(execute_btn, "Run the process and save the updated Excel file.")

        # Footer
        footer = tk.Label(master, text="© 2025 Excel Merge Tool | Help: Use standard Excel files with 'id' and 'datetime' columns.", font=("Segoe UI", 9), bg="#e8eaf6", fg="#4a4a4a", bd=1, relief="flat")
        footer.pack(side=tk.BOTTOM, fill=tk.X, pady=(15,0))

    def browse_file1(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.file1_path.set(path)



    def get_processed_dataframe(self):
        file1 = self.file1_path.get()
        dt_str = self.datetime_str.get()

        if not file1 or not dt_str:
            return None, "Please select the file and enter a date-time."
        try:
            dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
        except ValueError:
            return None, "Invalid date-time format. Use YYYY-MM-DD HH:MM"
        # --- New logic to preserve Excel formatting using openpyxl ---
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table
        import tempfile

        try:
            wb = openpyxl.load_workbook(file1)
            ws = wb.active
        except Exception as e:
            return None, f"Failed to read Excel file: {e}"

        # Find the header row and the 'End Time' column index
        header_row = None
        end_time_col_idx = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5), 1):
            for cell in row:
                if cell.value and str(cell.value).strip().lower() == 'end time':
                    header_row = i
                    end_time_col_idx = cell.col_idx
                    break
            if header_row:
                break
        if not header_row or not end_time_col_idx:
            return None, "File 1 must have an 'End Time' column."

        # Parse the filter datetime
        import datetime as dtmod
        try:
            filter_dt = dtmod.datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
        except Exception:
            return None, "Invalid date-time format. Use YYYY-MM-DD HH:MM"

        # Collect rows to keep (header + rows with End Time > filter_dt)
        rows_to_keep = []
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), 1):
            if idx <= header_row:
                rows_to_keep.append([cell for cell in row])
            else:
                cell = row[end_time_col_idx-1]
                try:
                    cell_dt = pd.to_datetime(cell.value, errors='coerce')
                except Exception:
                    cell_dt = None
                if cell_dt is not pd.NaT and pd.notnull(cell_dt) and cell_dt > filter_dt:
                    rows_to_keep.append([cell for cell in row])

        # Create a new workbook and copy rows with formatting
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        import copy
        for r_idx, row in enumerate(rows_to_keep, 1):
            for c_idx, cell in enumerate(row, 1):
                new_cell = new_ws.cell(row=r_idx, column=c_idx, value=cell.value)
                # Copy only public style attributes with copy.copy() and error handling
                if cell.has_style:
                    try:
                        if cell.font: new_cell.font = copy.copy(cell.font)
                    except Exception: pass
                    try:
                        if cell.fill: new_cell.fill = copy.copy(cell.fill)
                    except Exception: pass
                    try:
                        if cell.border: new_cell.border = copy.copy(cell.border)
                    except Exception: pass
                    try:
                        if cell.alignment: new_cell.alignment = copy.copy(cell.alignment)
                    except Exception: pass
                    try:
                        if cell.number_format: new_cell.number_format = cell.number_format
                    except Exception: pass
                    try:
                        if cell.protection: new_cell.protection = copy.copy(cell.protection)
                    except Exception: pass
                if cell.hyperlink:
                    new_cell.hyperlink = cell.hyperlink
                if cell.comment:
                    new_cell.comment = cell.comment
        # Copy column widths
        for col in ws.column_dimensions:
            new_ws.column_dimensions[col].width = ws.column_dimensions[col].width
        # Copy merged cells
        for merged_range in ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))
        # Copy tables if present
        if hasattr(ws, 'tables'):
            for tname, table in ws.tables.items():
                new_ws.add_table(Table(displayName=table.displayName, ref=table.ref))

        # Save to a temporary file and reload as DataFrame for preview
        tmpfile = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        new_wb.save(tmpfile.name)
        tmpfile.close()
        # For preview, load as DataFrame
        result_df = pd.read_excel(tmpfile.name, header=header_row-1)
        return result_df, None, tmpfile.name

    def process_preview(self):
        result_df, err = self.get_processed_dataframe()
        if err:
            messagebox.showerror("Preview Error", err)
            return
        msg = f"Preview:\nRows in result: {len(result_df)}\nColumns: {list(result_df.columns)}"
        messagebox.showinfo("Process Preview", msg)

    def process_files(self):
        result = self.get_processed_dataframe()
        # Now get_processed_dataframe returns (result_df, err, temp_file_path)
        if len(result) == 2:
            result_df, err = result
            temp_file_path = None
        else:
            result_df, err, temp_file_path = result
        if err:
            messagebox.showerror("Error", err)
            return
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Save updated File 1 as...")
        if save_path:
            try:
                if temp_file_path:
                    import shutil
                    shutil.copy(temp_file_path, save_path)
                else:
                    result_df.to_excel(save_path, index=False)
                messagebox.showinfo("Success", f"File saved to {save_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save file: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelMergeApp(root)
    root.mainloop()
